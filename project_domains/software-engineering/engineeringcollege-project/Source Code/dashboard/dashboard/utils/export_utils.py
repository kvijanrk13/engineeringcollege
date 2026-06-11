# dashboard/utils/export_utils.py
import csv
from io import StringIO, BytesIO
import logging

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Excel export will use CSV fallback.")


def export_to_excel(data, headers, filename=None):
    """Export data to Excel format"""
    if HAS_OPENPYXL:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Add data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            if filename:
                with open(filename, 'wb') as f:
                    f.write(output.getvalue())
                return filename
            return output

        except Exception as e:
            logger.error(f"Excel export error: {e}")
            # Fall back to CSV

    # CSV fallback
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(data)

    if filename:
        with open(filename.replace('.xlsx', '.csv'), 'w', newline='') as f:
            f.write(output.getvalue())
        return filename.replace('.xlsx', '.csv')

    return BytesIO(output.getvalue().encode())