# dashboard/utils.py
from datetime import date
import re
import tempfile
import os
import logging
from io import BytesIO, StringIO
import csv

logger = logging.getLogger(__name__)


# ==================== HELPER FUNCTIONS ====================

def calculate_experience(joining_date):
    """Calculate years of experience from joining date"""
    if not joining_date:
        return "N/A"
    today = date.today()
    years = today.year - joining_date.year
    months = today.month - joining_date.month
    days = today.day - joining_date.day

    if days < 0:
        months -= 1
        # Get days in previous month
        if today.month == 1:
            prev_month = 12
            year_for_days = today.year - 1
        else:
            prev_month = today.month - 1
            year_for_days = today.year

        if prev_month in [4, 6, 9, 11]:
            days_in_prev_month = 30
        elif prev_month == 2:
            # Check for leap year
            if (year_for_days % 4 == 0 and year_for_days % 100 != 0) or (year_for_days % 400 == 0):
                days_in_prev_month = 29
            else:
                days_in_prev_month = 28
        else:
            days_in_prev_month = 31

        days += days_in_prev_month

    if months < 0:
        years -= 1
        months += 12

    return f"{years} Years {months} Months"


def calculate_age(dob):
    """Calculate age from date of birth"""
    if not dob:
        return None
    today = date.today()
    age = today.year - dob.year
    # Adjust if birthday hasn't occurred yet this year
    if (today.month, today.day) < (dob.month, dob.day):
        age -= 1
    return age


def format_date(date_obj, format_str="%d-%m-%Y"):
    """Format date object to string"""
    if not date_obj:
        return ""
    return date_obj.strftime(format_str)


def get_academic_year(date_obj=None):
    """Get academic year for a given date"""
    if not date_obj:
        date_obj = date.today()
    year = date_obj.year
    month = date_obj.month
    if month >= 6:  # June is start of academic year
        return f"{year}-{year + 1}"
    else:
        return f"{year - 1}-{year}"


# ==================== PDF UTILITIES ====================

def generate_pdf_from_html(html_string, output_path=None):
    """Generate PDF from HTML string using WeasyPrint"""
    try:
        from weasyprint import HTML
        from django.conf import settings as django_settings

        base_url = f"file:///{django_settings.BASE_DIR}" if django_settings.BASE_DIR else None
        html_obj = HTML(string=html_string, base_url=base_url)

        if output_path:
            html_obj.write_pdf(target=output_path)
            return output_path
        else:
            pdf_bytes = html_obj.write_pdf()
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(pdf_bytes)
                return tmp.name
    except ImportError:
        # Fallback to reportlab if WeasyPrint not available
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet

        if not output_path:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            output_path = tmp.name
            tmp.close()

        doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = [Paragraph(html_string.replace('\n', '<br/>'), styles['Normal'])]
        doc.build(story)
        return output_path


def merge_pdfs(pdf_list, output_path=None):
    """Merge multiple PDF files"""
    from pypdf import PdfMerger

    merger = PdfMerger()
    temp_files = []

    for pdf in pdf_list:
        if pdf:
            merger.append(pdf)

    if output_path:
        merger.write(output_path)
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            merger.write(tmp.name)
            output_path = tmp.name
            temp_files.append(output_path)

    merger.close()
    return output_path, temp_files


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF"""
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        return f"Error extracting text: {str(e)}"


# ==================== VALIDATION FUNCTIONS ====================

def validate_faculty_data(data):
    """Validate faculty data"""
    errors = []
    warnings = []

    # Required fields
    required_fields = ['employee_code', 'staff_name', 'email', 'department']
    for field in required_fields:
        if not data.get(field):
            errors.append(f"{field} is required")

    # Email validation
    email = data.get('email', '')
    if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        errors.append("Invalid email format")

    # Phone validation
    mobile = data.get('mobile', '')
    if mobile and not re.match(r'^[0-9]{10}$', mobile):
        warnings.append("Mobile number should be 10 digits")

    # Date validations
    dob = data.get('dob')
    joining_date = data.get('joining_date')
    if dob and joining_date and dob > joining_date:
        warnings.append("Date of birth cannot be after joining date")

    return len(errors) == 0, errors, warnings


def validate_student_data(data):
    """Validate student data"""
    errors = []
    warnings = []

    # Required fields
    required_fields = ['ht_no', 'student_name']
    for field in required_fields:
        if not data.get(field):
            errors.append(f"{field} is required")

    # Email validation
    email = data.get('email', '')
    if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        errors.append("Invalid email format")

    # Phone validation
    student_phone = data.get('student_phone', '')
    if student_phone and not re.match(r'^[0-9]{10}$', student_phone):
        warnings.append("Student phone number should be 10 digits")

    # Roll number format
    ht_no = data.get('ht_no', '')
    if ht_no and not re.match(r'^[0-9A-Za-z]+$', ht_no):
        warnings.append("Hall ticket number should be alphanumeric")

    return len(errors) == 0, errors, warnings


# ==================== NOTIFICATION FUNCTIONS ====================

def send_email_notification(recipient, subject, message, from_email=None):
    """Send email notification"""
    from django.core.mail import send_mail
    from django.conf import settings

    try:
        if not from_email:
            from_email = settings.DEFAULT_FROM_EMAIL if hasattr(settings,
                                                                'DEFAULT_FROM_EMAIL') else 'webmaster@localhost'

        send_mail(
            subject,
            message,
            from_email,
            [recipient],
            fail_silently=False,
        )
        logger.info(f"Email sent to {recipient}: {subject}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email to {recipient}: {e}")
        return False


# ==================== QR CODE FUNCTIONS ====================

def generate_qr_code(data, size=10, border=4):
    """Generate QR code image"""
    try:
        import qrcode
        from django.core.files.base import ContentFile

        qr = qrcode.QRCode(
            version=1,
            box_size=size,
            border=border,
        )
        qr.add_data(data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convert to bytes
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        return buffer
    except ImportError:
        logger.warning("qrcode library not installed")
        return None
    except Exception as e:
        logger.error(f"Error generating QR code: {e}")
        return None


def generate_qr_code_base64(data, size=10):
    """Generate QR code and return as base64 string"""
    import base64
    buffer = generate_qr_code(data, size)
    if buffer:
        return base64.b64encode(buffer.getvalue()).decode()
    return None


def generate_qr_code_file(data, filename=None):
    """Generate QR code and return as Django ContentFile"""
    from django.core.files.base import ContentFile
    import hashlib

    buffer = generate_qr_code(data)
    if buffer:
        if not filename:
            filename = f"qr_{hashlib.md5(data.encode()).hexdigest()[:10]}.png"
        return ContentFile(buffer.getvalue(), name=filename)
    return None


# ==================== EXPORT FUNCTIONS ====================

def export_to_excel(data, headers, filename=None):
    """Export data to Excel format"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        if filename:
            with open(filename, 'wb') as f:
                f.write(output.getvalue())
            return filename
        return output

    except ImportError:
        logger.warning("openpyxl not installed. Using CSV fallback.")
        # CSV fallback
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)

        if filename:
            with open(filename.replace('.xlsx', '.csv'), 'w', newline='') as f:
                f.write(output.getvalue())
            return filename.replace('.xlsx', '.csv')

        return BytesIO(output.getvalue().encode())
    except Exception as e:
        logger.error(f"Export error: {e}")
        return None