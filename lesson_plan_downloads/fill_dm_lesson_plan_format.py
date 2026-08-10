from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


SOURCE_FORMAT = Path(r"E:\IT DEPT\LESSON PLANS\2026-2027\LESSON PLAN FORMAT.xlsx")
SOURCE_DATA = Path(r"E:\IT DEPT\LESSON PLANS\2026-2027\R22 3-1 DM.xlsx")
OUTPUT = Path(r"F:\IT DEPT DJANGO PROJECT\engineeringcollege\lesson_plan_downloads\Filled LESSON PLAN FORMAT - R22 3-1 DM.xlsx")

FORMAT_SHEET = "3-1-IT"
DATA_SHEET = "Lesson Plan"
FIRST_OUTPUT_ROW = 5
LAST_TEMPLATE_DATA_ROW = 123


def read_lessons():
    sheet = load_workbook(SOURCE_DATA, data_only=True)[DATA_SHEET]
    lessons = []
    current_week = None

    for row in range(7, sheet.max_row + 1):
        date = sheet.cell(row, 2).value
        day = sheet.cell(row, 3).value
        week = sheet.cell(row, 4).value
        topic = sheet.cell(row, 6).value
        if not all((date, day, topic)):
            continue
        if week is not None:
            current_week = int(week)
        if current_week is None:
            raise ValueError(f"Missing week number at source row {row}")
        lessons.append({"date": date, "day": day, "week": current_week, "topic": topic})

    if len(lessons) != 70:
        raise ValueError(f"Expected 70 lessons, found {len(lessons)}")
    return lessons


def unmerge_data_area(sheet):
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row >= FIRST_OUTPUT_ROW and merged_range.min_row <= LAST_TEMPLATE_DATA_ROW:
            sheet.unmerge_cells(str(merged_range))


def apply_cell_style(cell, *, horizontal="center", wrap=False):
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical="center",
        wrap_text=wrap,
    )
    cell.font = Font(name="Arial", size=9)


def main():
    lessons = read_lessons()
    workbook = load_workbook(SOURCE_FORMAT)
    sheet = workbook[FORMAT_SHEET]

    unmerge_data_area(sheet)

    for row in range(FIRST_OUTPUT_ROW, LAST_TEMPLATE_DATA_ROW + 1):
        for column in range(1, 7):
            sheet.cell(row, column).value = None
        sheet.row_dimensions[row].hidden = row >= FIRST_OUTPUT_ROW + len(lessons)

    # Preserve the source file's institutional heading and format, but ensure the
    # subject label remains accurate.
    if "DATA MINING" not in str(sheet["A3"].value).upper():
        raise ValueError("The destination format does not identify the subject as Data Mining")

    for index, lesson in enumerate(lessons):
        row = FIRST_OUTPUT_ROW + index
        sheet.cell(row, 1, lesson["date"])
        sheet.cell(row, 2, lesson["day"])
        sheet.cell(row, 5, lesson["topic"])
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)

        apply_cell_style(sheet.cell(row, 1))
        apply_cell_style(sheet.cell(row, 2))
        apply_cell_style(sheet.cell(row, 5), horizontal="left", wrap=True)
        sheet.row_dimensions[row].height = 24 if len(str(lesson["topic"])) > 60 else 19

    # Create one merged Week No. and Classes Per Week cell for every actual
    # teaching week in the source workbook.
    start = 0
    while start < len(lessons):
        week = lessons[start]["week"]
        end = start
        while end + 1 < len(lessons) and lessons[end + 1]["week"] == week:
            end += 1

        first_row = FIRST_OUTPUT_ROW + start
        last_row = FIRST_OUTPUT_ROW + end
        class_count = end - start + 1
        sheet.merge_cells(start_row=first_row, start_column=3, end_row=last_row, end_column=3)
        sheet.merge_cells(start_row=first_row, start_column=4, end_row=last_row, end_column=4)
        sheet.cell(first_row, 3, week)
        sheet.cell(first_row, 4, class_count)
        apply_cell_style(sheet.cell(first_row, 3))
        apply_cell_style(sheet.cell(first_row, 4))
        start = end + 1

    # Hide obsolete rows belonging to the old academic calendar embedded in the
    # blank template, and print only the completed teaching plan.
    for row in range(FIRST_OUTPUT_ROW + len(lessons), sheet.max_row + 1):
        sheet.row_dimensions[row].hidden = True
    sheet.print_area = f"A1:F{FIRST_OUTPUT_ROW + len(lessons) - 1}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
