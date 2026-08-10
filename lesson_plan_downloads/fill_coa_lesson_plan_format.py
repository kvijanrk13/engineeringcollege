from pathlib import Path

from openpyxl import load_workbook


SOURCE_FORMAT = Path(
    r"F:\IT DEPT DJANGO PROJECT\engineeringcollege\lesson_plan_downloads\Filled LESSON PLAN FORMAT - R22 3-1 DM.xlsx"
)
SOURCE_DATA = Path(r"E:\IT DEPT\LESSON PLANS\2026-2027\R25 2-1 COA.xlsx")
OUTPUT = Path(
    r"F:\IT DEPT DJANGO PROJECT\engineeringcollege\lesson_plan_downloads\Filled LESSON PLAN FORMAT - R25 2-1 COA.xlsx"
)


def main():
    data_sheet = load_workbook(SOURCE_DATA, data_only=True)["Lesson Plan"]
    lessons = []
    current_week = None

    for row in range(7, data_sheet.max_row + 1):
        date = data_sheet.cell(row, 2).value
        day = data_sheet.cell(row, 3).value
        week = data_sheet.cell(row, 4).value
        topic = data_sheet.cell(row, 6).value
        if not all((date, day, topic)):
            continue
        if week is not None:
            current_week = int(week)
        lessons.append((date, day, current_week, topic))

    if len(lessons) != 70:
        raise ValueError(f"Expected 70 lessons, found {len(lessons)}")

    workbook = load_workbook(SOURCE_FORMAT)
    sheet = workbook["3-1-IT"]
    heading = str(sheet["A3"].value)
    sheet["A3"] = heading.replace("DATA MINING", "COMPUTER ORGANIZATION")

    for index, (date, day, _week, topic) in enumerate(lessons, start=5):
        sheet.cell(index, 1, date)
        sheet.cell(index, 2, day)
        sheet.cell(index, 5, topic)
        sheet.row_dimensions[index].height = 24 if len(str(topic)) > 60 else 19

    # The COA source uses the same dates and weekly grouping as the prepared
    # format. Verify every merged week block before saving.
    expected_week_counts = []
    for _date, _day, week, _topic in lessons:
        if not expected_week_counts or expected_week_counts[-1][0] != week:
            expected_week_counts.append([week, 1])
        else:
            expected_week_counts[-1][1] += 1

    actual_week_counts = [
        (sheet.cell(row, 3).value, sheet.cell(row, 4).value)
        for row in range(5, 75)
        if sheet.cell(row, 3).value is not None
    ]
    if [tuple(item) for item in expected_week_counts] != actual_week_counts:
        raise ValueError("COA weekly grouping does not match the destination format")

    workbook.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
