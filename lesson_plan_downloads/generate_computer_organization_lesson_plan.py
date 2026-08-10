from pathlib import Path
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
SOURCE = BASE_DIR / "lesson_plan_merged_weeks.xlsx"
OUTPUT = BASE_DIR / "computer_organization_lesson_plan_mid_exam_revised.xlsx"
MID_EXAM_DATE = "29-08-2026"


TOPICS = [
    # Unit I: periods 1-18
    "UNIT-I - Data Representation: Number Systems and Positional Notation",
    "Number Systems: Decimal, Binary, Octal and Hexadecimal Conversions",
    "Complements: Radix and Diminished Radix Complements",
    "Arithmetic Addition in Binary, Octal and Hexadecimal Systems",
    "Arithmetic Subtraction Using Complements and Overflow Detection",
    "Digital Logic Circuits: Basic and Universal Logic Gates",
    "Boolean Algebra: Postulates, Laws, Theorems and Simplification",
    "Map Simplification: Two-Variable and Three-Variable Karnaugh Maps",
    "Map Simplification: Four-Variable Maps and Don't-Care Conditions",
    "Combinational Circuits: Half Adder and Full Adder",
    "Flip-Flops: SR and D Flip-Flops",
    "Flip-Flops: JK and T Flip-Flops",
    "Sequential Circuits: Flip-Flop Input and Characteristic Equations",
    "Sequential Circuits: State Table and State Diagram",
    "Digital Components: Decoders and Their Applications",
    "Digital Components: Multiplexers and Their Applications",
    "Digital Components: Registers and Binary Counters",
    "Digital Computers: Block Diagram; Computer Organization, Design and Architecture",
    # Unit II: periods 19-32
    "UNIT-II - Register Transfer Language: Symbols, Statements and Control Functions",
    "Register Transfer and Conditional Register Transfer",
    "Bus Transfers and Common Bus Organization",
    "Memory Read and Memory Write Transfers",
    "Arithmetic Micro Operations",
    "Logic Micro Operations",
    "Shift Micro Operations",
    "Arithmetic Logic Shift Unit",
    "Basic Computer Organization: Instruction Codes and Stored Program Concept",
    "Computer Registers, Computer Instructions and Common Bus System",
    "Timing and Control: Control Unit and Timing Signals",
    "Instruction Cycle: Fetch, Decode, Execute and Interrupt Phases",
    "Memory-Reference Instructions",
    "Input-Output Instructions and Interrupt Processing",
    # First half of Unit III: periods 33-38, completed by 29-08-2026
    "UNIT-III (First Half) - Microprogrammed Control: Control Memory and Microinstructions",
    "Address Sequencing in a Microprogrammed Control Unit",
    "Conditional Branching, Mapping and Microprogram Subroutines",
    "Microprogram Example for a Basic Computer",
    "Design of a Microprogrammed Control Unit",
    "Central Processing Unit: General Register Organization",
    # Remaining half of Unit III: periods 39-44
    "UNIT-III (Second Half) - Instruction Formats",
    "Addressing Modes: Immediate, Direct, Indirect and Register Modes",
    "Addressing Modes: Register Indirect, Relative, Indexed and Base Register Modes",
    "Data Transfer Instructions",
    "Data Manipulation: Arithmetic, Logical and Shift Instructions",
    "Program Control: Branch, Call, Return and Status Conditions",
    # Unit IV: periods 45-58
    "UNIT-IV - Input-Output Organization: I/O Interface and I/O Bus",
    "Asynchronous Data Transfer: Concepts and Requirements",
    "Asynchronous Data Transfer: Strobe Control",
    "Asynchronous Data Transfer: Handshaking",
    "Modes of Transfer: Programmed I/O and Interrupt-Initiated I/O",
    "Priority Interrupt: Daisy Chaining and Parallel Priority",
    "Direct Memory Access: DMA Controller and DMA Transfer",
    "DMA Transfer Modes and CPU-DMA Bus Arbitration",
    "Memory Organization: Memory Hierarchy and Performance Parameters",
    "Main Memory: RAM, ROM and Memory Address Map",
    "Main Memory Organization, Expansion and Address Decoding",
    "Auxiliary Memory: Magnetic Disk, Tape and Optical Storage",
    "Associative Memory: Hardware Organization and Match Logic",
    "Cache Memory: Locality, Hit Ratio and Cache Mapping Techniques",
    # Unit V: periods 59-70
    "UNIT-V - Pipeline and Vector Processing: Parallel Processing Concepts",
    "Parallel Processing: Classification, Performance and Speedup",
    "Pipelining: Principles and Pipeline Performance",
    "Arithmetic Pipeline",
    "Instruction Pipeline",
    "Instruction Pipeline Hazards and Their Handling",
    "Vector Processing and Vector Operations",
    "Vector Processor Organization",
    "Multiprocessors: Characteristics and Classification",
    "Interconnection Structures: Time-Shared Bus and Crossbar Switch",
    "Interconnection Structures: Multistage Networks and Hypercube",
    "Characteristics and Comparison of RISC and CISC",
]


def main():
    if len(TOPICS) != 70:
        raise ValueError(f"Expected 70 lesson topics, found {len(TOPICS)}")

    workbook = load_workbook(SOURCE)
    sheet = workbook["Lesson Plan"]
    sheet["D4"] = "Computer Organization"

    for row_number, topic in enumerate(TOPICS, start=7):
        sheet.cell(row=row_number, column=6, value=topic)

    cutoff_row = next(
        row for row in range(7, 77) if sheet.cell(row, 2).value == MID_EXAM_DATE
    )
    if cutoff_row != 44 or "General Register Organization" not in sheet["F44"].value:
        raise ValueError("The 2.5-unit mid-examination cutoff is not aligned to 29-08-2026")

    workbook.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
